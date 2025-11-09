from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages
from django.db.models import Q, Count
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
from django.core.mail import send_mail
from django.conf import settings
import json

from .models import Espacio, Reserva, TipoEspacio, PerfilUsuario
from .forms import (
    ReservaForm, EspacioForm, TipoEspacioForm, 
    PerfilUsuarioForm, UserRegistrationForm
)

# Importaciones para reportes
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

# Vista pública - Página de inicio
def home(request):
    espacios_destacados = Espacio.objects.filter(activo=True)[:6]
    tipos_espacios = TipoEspacio.objects.all()
    context = {
        'espacios_destacados': espacios_destacados,
        'tipos_espacios': tipos_espacios,
    }
    return render(request, 'reservas/home.html', context)

# Registro de usuarios
# Registro de usuarios
def registro(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Crear perfil automáticamente
            PerfilUsuario.objects.create(
                user=user,
                rol='USUARIO',
                telefono=form.cleaned_data.get('telefono', ''),
                departamento=form.cleaned_data.get('departamento', '')
            )
            messages.success(request, '¡Registro exitoso! Ya puedes iniciar sesión con tu usuario y contraseña.')
            # Redirigir al login en lugar de auto-login
            return redirect('login')
    else:
        form = UserRegistrationForm()
    return render(request, 'registration/registro.html', {'form': form})

# Dashboard principal
@login_required
def dashboard(request):
    user = request.user
    es_admin = hasattr(user, 'perfil') and user.perfil.rol == 'ADMINISTRADOR'
    
    if es_admin:
        # Estadísticas para administrador
        total_espacios = Espacio.objects.filter(activo=True).count()
        total_reservas = Reserva.objects.count()
        reservas_pendientes = Reserva.objects.filter(estado='PENDIENTE').count()
        reservas_hoy = Reserva.objects.filter(fecha_reserva=timezone.now().date()).count()
        
        # Reservas recientes
        reservas_recientes = Reserva.objects.select_related('usuario', 'espacio').order_by('-fecha_creacion')[:10]
        
        context = {
            'es_admin': es_admin,
            'total_espacios': total_espacios,
            'total_reservas': total_reservas,
            'reservas_pendientes': reservas_pendientes,
            'reservas_hoy': reservas_hoy,
            'reservas_recientes': reservas_recientes,
        }
    else:
        # Dashboard para usuario normal
        mis_reservas = Reserva.objects.filter(usuario=user).order_by('-fecha_reserva')[:5]
        proximas_reservas = Reserva.objects.filter(
            usuario=user,
            fecha_reserva__gte=timezone.now().date(),
            estado__in=['PENDIENTE', 'CONFIRMADA']
        ).order_by('fecha_reserva', 'hora_inicio')[:5]
        
        context = {
            'es_admin': es_admin,
            'mis_reservas': mis_reservas,
            'proximas_reservas': proximas_reservas,
        }
    
    return render(request, 'reservas/dashboard.html', context)

# CRUD de Espacios
@login_required
def lista_espacios(request):
    espacios = Espacio.objects.filter(activo=True).select_related('tipo')
    
    # Filtros
    tipo_filtro = request.GET.get('tipo')
    busqueda = request.GET.get('q')
    
    if tipo_filtro:
        espacios = espacios.filter(tipo_id=tipo_filtro)
    
    if busqueda:
        espacios = espacios.filter(
            Q(nombre__icontains=busqueda) |
            Q(codigo__icontains=busqueda) |
            Q(ubicacion__icontains=busqueda)
        )
    
    tipos_espacios = TipoEspacio.objects.all()
    
    context = {
        'espacios': espacios,
        'tipos_espacios': tipos_espacios,
        'tipo_seleccionado': tipo_filtro,
        'busqueda': busqueda,
    }
    return render(request, 'reservas/espacios_lista.html', context)

@login_required
def detalle_espacio(request, pk):
    espacio = get_object_or_404(Espacio, pk=pk)
    reservas_futuras = Reserva.objects.filter(
        espacio=espacio,
        fecha_reserva__gte=timezone.now().date(),
        estado__in=['CONFIRMADA', 'PENDIENTE']
    ).order_by('fecha_reserva', 'hora_inicio')[:5]
    
    context = {
        'espacio': espacio,
        'reservas_futuras': reservas_futuras,
    }
    return render(request, 'reservas/espacio_detalle.html', context)

@login_required
def crear_espacio(request):
    if not (hasattr(request.user, 'perfil') and request.user.perfil.rol == 'ADMINISTRADOR'):
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return redirect('lista_espacios')
    
    if request.method == 'POST':
        form = EspacioForm(request.POST, request.FILES)
        if form.is_valid():
            espacio = form.save()
            messages.success(request, f'Espacio "{espacio.nombre}" creado exitosamente.')
            return redirect('detalle_espacio', pk=espacio.pk)
    else:
        form = EspacioForm()
    
    return render(request, 'reservas/espacio_form.html', {'form': form, 'accion': 'Crear'})

@login_required
def editar_espacio(request, pk):
    if not (hasattr(request.user, 'perfil') and request.user.perfil.rol == 'ADMINISTRADOR'):
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return redirect('lista_espacios')
    
    espacio = get_object_or_404(Espacio, pk=pk)
    
    if request.method == 'POST':
        form = EspacioForm(request.POST, request.FILES, instance=espacio)
        if form.is_valid():
            form.save()
            messages.success(request, f'Espacio "{espacio.nombre}" actualizado exitosamente.')
            return redirect('detalle_espacio', pk=espacio.pk)
    else:
        form = EspacioForm(instance=espacio)
    
    return render(request, 'reservas/espacio_form.html', {'form': form, 'accion': 'Editar', 'espacio': espacio})

@login_required
def eliminar_espacio(request, pk):
    if not (hasattr(request.user, 'perfil') and request.user.perfil.rol == 'ADMINISTRADOR'):
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return redirect('lista_espacios')
    
    espacio = get_object_or_404(Espacio, pk=pk)
    
    if request.method == 'POST':
        nombre = espacio.nombre
        espacio.activo = False
        espacio.save()
        messages.success(request, f'Espacio "{nombre}" desactivado exitosamente.')
        return redirect('lista_espacios')
    
    return render(request, 'reservas/espacio_confirmar_eliminacion.html', {'espacio': espacio})

# CRUD de Reservas
@login_required
def lista_reservas(request):
    user = request.user
    es_admin = hasattr(user, 'perfil') and user.perfil.rol == 'ADMINISTRADOR'
    
    if es_admin:
        reservas = Reserva.objects.all().select_related('usuario', 'espacio')
    else:
        reservas = Reserva.objects.filter(usuario=user).select_related('espacio')
    
    # Filtros
    estado_filtro = request.GET.get('estado')
    fecha_filtro = request.GET.get('fecha')
    
    if estado_filtro:
        reservas = reservas.filter(estado=estado_filtro)
    
    if fecha_filtro:
        reservas = reservas.filter(fecha_reserva=fecha_filtro)
    
    reservas = reservas.order_by('-fecha_reserva', '-hora_inicio')
    
    context = {
        'reservas': reservas,
        'es_admin': es_admin,
        'estado_seleccionado': estado_filtro,
        'fecha_seleccionada': fecha_filtro,
    }
    return render(request, 'reservas/reservas_lista.html', context)

@login_required
def detalle_reserva(request, pk):
    reserva = get_object_or_404(Reserva, pk=pk)
    
    # Verificar permisos
    es_admin = hasattr(request.user, 'perfil') and request.user.perfil.rol == 'ADMINISTRADOR'
    if not es_admin and reserva.usuario != request.user:
        messages.error(request, 'No tienes permisos para ver esta reserva.')
        return redirect('lista_reservas')
    
    context = {
        'reserva': reserva,
        'es_admin': es_admin,
    }
    return render(request, 'reservas/reserva_detalle.html', context)

@login_required
def crear_reserva(request):
    if request.method == 'POST':
        form = ReservaForm(request.POST)
        if form.is_valid():
            reserva = form.save(commit=False)
            reserva.usuario = request.user
            try:
                reserva.save()
                messages.success(request, 'Reserva creada exitosamente.')
                
                # Enviar email de recordatorio (simulado en consola)
                if hasattr(request.user, 'perfil') and request.user.perfil.notificaciones_email:
                    send_mail(
                        f'Confirmación de Reserva - {reserva.espacio.nombre}',
                        f'Tu reserva para {reserva.fecha_reserva} de {reserva.hora_inicio} a {reserva.hora_fin} ha sido {reserva.estado.lower()}.',
                        settings.DEFAULT_FROM_EMAIL,
                        [request.user.email],
                        fail_silently=True,
                    )
                
                return redirect('detalle_reserva', pk=reserva.pk)
            except Exception as e:
                messages.error(request, f'Error al crear la reserva: {str(e)}')
    else:
        # Pre-seleccionar espacio si viene en URL
        espacio_id = request.GET.get('espacio')
        initial = {}
        if espacio_id:
            initial['espacio'] = espacio_id
        form = ReservaForm(initial=initial)
    
    return render(request, 'reservas/reserva_form.html', {'form': form, 'accion': 'Crear'})

@login_required
def editar_reserva(request, pk):
    reserva = get_object_or_404(Reserva, pk=pk)
    
    # Verificar permisos
    es_admin = hasattr(request.user, 'perfil') and request.user.perfil.rol == 'ADMINISTRADOR'
    if not es_admin and reserva.usuario != request.user:
        messages.error(request, 'No tienes permisos para editar esta reserva.')
        return redirect('lista_reservas')
    
    # No permitir editar reservas pasadas o canceladas
    if reserva.estado == 'CANCELADA' or reserva.fecha_reserva < timezone.now().date():
        messages.error(request, 'No se puede editar una reserva cancelada o pasada.')
        return redirect('detalle_reserva', pk=pk)
    
    if request.method == 'POST':
        form = ReservaForm(request.POST, instance=reserva)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Reserva actualizada exitosamente.')
                return redirect('detalle_reserva', pk=reserva.pk)
            except Exception as e:
                messages.error(request, f'Error al actualizar la reserva: {str(e)}')
    else:
        form = ReservaForm(instance=reserva)
    
    return render(request, 'reservas/reserva_form.html', {'form': form, 'accion': 'Editar', 'reserva': reserva})

@login_required
def cancelar_reserva(request, pk):
    reserva = get_object_or_404(Reserva, pk=pk)
    
    # Verificar permisos
    es_admin = hasattr(request.user, 'perfil') and request.user.perfil.rol == 'ADMINISTRADOR'
    if not es_admin and reserva.usuario != request.user:
        messages.error(request, 'No tienes permisos para cancelar esta reserva.')
        return redirect('lista_reservas')
    
    if request.method == 'POST':
        reserva.estado = 'CANCELADA'
        reserva.save()
        messages.success(request, 'Reserva cancelada exitosamente.')
        return redirect('lista_reservas')
    
    return render(request, 'reservas/reserva_confirmar_cancelacion.html', {'reserva': reserva})

@login_required
def confirmar_reserva(request, pk):
    # Solo administradores pueden confirmar
    if not (hasattr(request.user, 'perfil') and request.user.perfil.rol == 'ADMINISTRADOR'):
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return redirect('lista_reservas')
    
    reserva = get_object_or_404(Reserva, pk=pk)
    
    if reserva.estado == 'PENDIENTE':
        reserva.estado = 'CONFIRMADA'
        reserva.confirmada_por = request.user
        reserva.save()
        messages.success(request, f'Reserva confirmada exitosamente.')
    else:
        messages.info(request, 'La reserva ya fue procesada.')
    
    return redirect('detalle_reserva', pk=pk)

# Reportes y estadísticas
@login_required
def reportes(request):
    if not (hasattr(request.user, 'perfil') and request.user.perfil.rol == 'ADMINISTRADOR'):
        messages.error(request, 'No tienes permisos para ver los reportes.')
        return redirect('dashboard')
    
    # Estadísticas generales
    total_espacios = Espacio.objects.filter(activo=True).count()
    total_reservas = Reserva.objects.count()
    total_usuarios = PerfilUsuario.objects.count()
    
    # Reservas por estado
    reservas_por_estado = Reserva.objects.values('estado').annotate(total=Count('id'))
    
    # Espacios más reservados
    espacios_populares = Reserva.objects.values('espacio__nombre').annotate(
        total=Count('id')
    ).order_by('-total')[:5]
    
    # Reservas por tipo de espacio
    reservas_por_tipo = Reserva.objects.values('espacio__tipo__nombre').annotate(
        total=Count('id')
    ).order_by('-total')
    
    # Ocupación semanal (últimos 7 días)
    hoy = timezone.now().date()
    hace_7_dias = hoy - timedelta(days=6)
    ocupacion_semanal = []
    
    for i in range(7):
        fecha = hace_7_dias + timedelta(days=i)
        total = Reserva.objects.filter(
            fecha_reserva=fecha,
            estado__in=['CONFIRMADA', 'COMPLETADA']
        ).count()
        ocupacion_semanal.append({
            'fecha': fecha.strftime('%d/%m'),
            'total': total
        })
    
    context = {
        'total_espacios': total_espacios,
        'total_reservas': total_reservas,
        'total_usuarios': total_usuarios,
        'reservas_por_estado': list(reservas_por_estado),
        'espacios_populares': list(espacios_populares),
        'reservas_por_tipo': list(reservas_por_tipo),
        'ocupacion_semanal': ocupacion_semanal,
    }
    
    return render(request, 'reservas/reportes.html', context)

# API para calendario
@login_required
def api_reservas_calendario(request):
    espacio_id = request.GET.get('espacio')
    
    if not espacio_id:
        return JsonResponse({'error': 'Se requiere ID de espacio'}, status=400)
    
    reservas = Reserva.objects.filter(
        espacio_id=espacio_id,
        estado__in=['CONFIRMADA', 'PENDIENTE']
    )
    
    eventos = []
    for reserva in reservas:
        eventos.append({
            'id': reserva.id,
            'title': f'{reserva.usuario.get_full_name() or reserva.usuario.username} - {reserva.motivo}',
            'start': f'{reserva.fecha_reserva}T{reserva.hora_inicio}',
            'end': f'{reserva.fecha_reserva}T{reserva.hora_fin}',
            'backgroundColor': '#28a745' if reserva.estado == 'CONFIRMADA' else '#ffc107',
            'borderColor': '#28a745' if reserva.estado == 'CONFIRMADA' else '#ffc107',
        })
    
    return JsonResponse(eventos, safe=False)

# Historial de reservas del usuario
@login_required
def historial_reservas(request):
    reservas = Reserva.objects.filter(usuario=request.user).select_related('espacio').order_by('-fecha_reserva')
    
    # Filtros
    estado_filtro = request.GET.get('estado')
    if estado_filtro:
        reservas = reservas.filter(estado=estado_filtro)
    
    context = {
        'reservas': reservas,
        'estado_seleccionado': estado_filtro,
    }
    return render(request, 'reservas/historial_reservas.html', context)

# Perfil de usuario
@login_required
def perfil_usuario(request):
    perfil, created = PerfilUsuario.objects.get_or_create(user=request.user)
    
    if request.method == 'POST':
        form = PerfilUsuarioForm(request.POST, instance=perfil)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil actualizado exitosamente.')
            return redirect('perfil_usuario')
    else:
        form = PerfilUsuarioForm(instance=perfil)
    
    # Estadísticas del usuario
    total_reservas = Reserva.objects.filter(usuario=request.user).count()
    reservas_activas = Reserva.objects.filter(
        usuario=request.user,
        estado__in=['CONFIRMADA', 'PENDIENTE'],
        fecha_reserva__gte=timezone.now().date()
    ).count()
    
    context = {
        'form': form,
        'perfil': perfil,
        'total_reservas': total_reservas,
        'reservas_activas': reservas_activas,
    }
    return render(request, 'reservas/perfil_usuario.html', context)

# EXPORTACIÓN DE REPORTES
@login_required
def exportar_reporte_pdf(request):
    if not (hasattr(request.user, 'perfil') and request.user.perfil.rol == 'ADMINISTRADOR'):
        messages.error(request, 'No tienes permisos para exportar reportes.')
        return redirect('dashboard')
    
    # Crear buffer
    buffer = BytesIO()
    
    # Crear documento PDF
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elementos = []
    
    # Estilos
    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        'CustomTitle',
        parent=estilos['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#0d6efd'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Título
    titulo = Paragraph("Reporte de Reservas del Sistema", estilo_titulo)
    elementos.append(titulo)
    elementos.append(Spacer(1, 0.3*inch))
    
    # Fecha
    fecha_reporte = Paragraph(
        f"Fecha de generación: {timezone.now().strftime('%d/%m/%Y %H:%M')}",
        estilos['Normal']
    )
    elementos.append(fecha_reporte)
    elementos.append(Spacer(1, 0.3*inch))
    
    # Estadísticas generales
    total_espacios = Espacio.objects.filter(activo=True).count()
    total_reservas = Reserva.objects.count()
    reservas_confirmadas = Reserva.objects.filter(estado='CONFIRMADA').count()
    reservas_pendientes = Reserva.objects.filter(estado='PENDIENTE').count()
    
    datos_estadisticas = [
        ['Concepto', 'Cantidad'],
        ['Total Espacios Activos', str(total_espacios)],
        ['Total Reservas', str(total_reservas)],
        ['Reservas Confirmadas', str(reservas_confirmadas)],
        ['Reservas Pendientes', str(reservas_pendientes)],
    ]
    
    tabla_estadisticas = Table(datos_estadisticas, colWidths=[4*inch, 2*inch])
    tabla_estadisticas.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elementos.append(tabla_estadisticas)
    elementos.append(Spacer(1, 0.5*inch))
    
    # Últimas 20 reservas
    subtitulo = Paragraph("Últimas 20 Reservas", estilos['Heading2'])
    elementos.append(subtitulo)
    elementos.append(Spacer(1, 0.2*inch))
    
    reservas = Reserva.objects.select_related('usuario', 'espacio').order_by('-fecha_creacion')[:20]
    
    datos_reservas = [['Espacio', 'Usuario', 'Fecha', 'Estado']]
    
    for reserva in reservas:
        datos_reservas.append([
            reserva.espacio.codigo,
            reserva.usuario.username,
            reserva.fecha_reserva.strftime('%d/%m/%Y'),
            reserva.get_estado_display()
        ])
    
    tabla_reservas = Table(datos_reservas, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    tabla_reservas.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#198754')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elementos.append(tabla_reservas)
    
    # Construir PDF
    doc.build(elementos)
    
    # Obtener valor del buffer
    pdf = buffer.getvalue()
    buffer.close()
    
    # Crear respuesta
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_reservas_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
    response.write(pdf)
    
    return response

@login_required
def exportar_reporte_excel(request):
    if not (hasattr(request.user, 'perfil') and request.user.perfil.rol == 'ADMINISTRADOR'):
        messages.error(request, 'No tienes permisos para exportar reportes.')
        return redirect('dashboard')
    
    # Crear libro de Excel
    wb = Workbook()
    
    # Hoja 1: Estadísticas Generales
    ws1 = wb.active
    ws1.title = "Estadísticas"
    
    # Encabezados con estilo
    ws1['A1'] = 'Reporte de Reservas del Sistema'
    ws1['A1'].font = Font(size=16, bold=True)
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A1:B1')
    
    ws1['A3'] = f"Fecha de generación: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    
    # Estadísticas
    ws1['A5'] = 'Concepto'
    ws1['B5'] = 'Cantidad'
    ws1['A5'].font = Font(bold=True)
    ws1['B5'].font = Font(bold=True)
    ws1['A5'].fill = PatternFill(start_color='0d6efd', end_color='0d6efd', fill_type='solid')
    ws1['B5'].fill = PatternFill(start_color='0d6efd', end_color='0d6efd', fill_type='solid')
    
    estadisticas = [
        ['Total Espacios Activos', Espacio.objects.filter(activo=True).count()],
        ['Total Reservas', Reserva.objects.count()],
        ['Reservas Confirmadas', Reserva.objects.filter(estado='CONFIRMADA').count()],
        ['Reservas Pendientes', Reserva.objects.filter(estado='PENDIENTE').count()],
        ['Reservas Canceladas', Reserva.objects.filter(estado='CANCELADA').count()],
    ]
    
    fila = 6
    for concepto, cantidad in estadisticas:
        ws1[f'A{fila}'] = concepto
        ws1[f'B{fila}'] = cantidad
        fila += 1
    
    # Hoja 2: Todas las Reservas
    ws2 = wb.create_sheet(title="Reservas")
    
    encabezados = ['ID', 'Espacio', 'Código', 'Usuario', 'Fecha', 'Hora Inicio', 'Hora Fin', 'Motivo', 'Estado']
    ws2.append(encabezados)
    
    # Estilo para encabezados
    for cell in ws2[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    reservas = Reserva.objects.select_related('usuario', 'espacio').order_by('-fecha_reserva')
    
    for reserva in reservas:
        ws2.append([
            reserva.id,
            reserva.espacio.nombre,
            reserva.espacio.codigo,
            reserva.usuario.get_full_name() or reserva.usuario.username,
            reserva.fecha_reserva.strftime('%d/%m/%Y'),
            reserva.hora_inicio.strftime('%H:%M'),
            reserva.hora_fin.strftime('%H:%M'),
            reserva.motivo,
            reserva.get_estado_display()
        ])
    
    # Ajustar ancho de columnas
    for column in ws2.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws2.column_dimensions[column_letter].width = adjusted_width
    
    # Hoja 3: Espacios
    ws3 = wb.create_sheet(title="Espacios")
    
    encabezados_espacios = ['Código', 'Nombre', 'Tipo', 'Capacidad', 'Ubicación', 'Activo']
    ws3.append(encabezados_espacios)
    
    for cell in ws3[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='ffc107', end_color='ffc107', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    espacios = Espacio.objects.select_related('tipo').all()
    
    for espacio in espacios:
        ws3.append([
            espacio.codigo,
            espacio.nombre,
            espacio.tipo.nombre,
            espacio.capacidad,
            espacio.ubicacion,
            'Sí' if espacio.activo else 'No'
        ])
    
    # Ajustar ancho de columnas
    for column in ws3.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws3.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Crear respuesta
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_reservas_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    return response